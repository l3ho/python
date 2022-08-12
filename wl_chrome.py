from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
import openpyxl
import sys, os, shutil
import time
import ctypes


def get_config_values(config_path):
    with open(config_path) as f:
        lines = f.readlines()
    config_values = {}
    for txt_line in lines:
        tmp_val = txt_line.split("|")
        config_values[tmp_val[0]] = tmp_val[1].replace("\n", "")
    f.close()
    return config_values

def find_column_by_name(column_name, wb):
    column_id = 0
    for i in range(1, 20):
        if wb.worksheets[0].cell(1, i).value == column_name:
            column_id = i
            break
    return column_id

def find_last_column(wb):
    ii = 1
    while wb.worksheets[0].cell(1, ii).value is not None:
        ii += 1
    return ii-1

def split_acc(acc_number):
    acc_ar = []
    cntr = 0
    tmp_st = ""
    for i in range(len(acc_number)):
        tmp_st = tmp_st + acc_number[i]
        if i == 1:
            acc_ar.append(tmp_st)
            tmp_st = ""
            cntr = 0
        if cntr == 4:
            acc_ar.append(tmp_st)
            tmp_st = ""
            cntr = 0
        cntr += 1
    return acc_ar

def fix_nip(nip_number):
    new_nip = ""
    for nm in nip_number:
        if nm.isdigit():
            new_nip += nm
    return new_nip

def main():
    if getattr(sys, 'frozen', False):
        app_path = os.path.dirname(sys.executable)
    else:
        app_path = os.path.dirname(os.path.abspath(__file__))
    config_path = os.path.join(app_path, "config.txt")

    config_dict = get_config_values(config_path)
    excel_path = os.path.join(app_path, config_dict["Input_file_name"])

    wb = openpyxl.open(filename=excel_path)
    acc_column = find_column_by_name(config_dict["Bank_account_column"], wb)
    if acc_column == 0:
        ctypes.windll.user32.MessageBoxW(0, "Could not find the account column - " + config_dict["Bank_account_column"], "Error", 1)
        exit()
    search_column = find_column_by_name("Search ID", wb)
    if search_column == 0:
        last_col = find_last_column(wb)
        wb.worksheets[0].cell(1, last_col + 1).value = "Search ID"
        wb.worksheets[0].cell(1, last_col + 2).value = "Status"
        search_column = last_col + 1

    driver_path = os.path.join(app_path, "chromedriver.exe")
    driver = webdriver.Chrome('chromedriver.exe')
    driver.get("https://www.podatki.gov.pl/wykaz-podatnikow-vat-wyszukiwarka")
    first_search = True

    ii = 2
    while wb.worksheets[0].cell(ii, acc_column).value != None:
        if wb.worksheets[0].cell(ii, search_column).value is None and wb.worksheets[0].cell(ii, acc_column).value != "Bank account cannot be found":
            search_id = ""
            item_found = False
            status_msg = ""

            acc_input_val = str(wb.worksheets[0].cell(ii, acc_column).value)
            acc_input_val = acc_input_val.replace("PL", "")
            if config_dict["Run_mode"] == "vat":
                vat_radio = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "//*[@id='wyszukiwarka']/div[1]/div[1]/fieldset[2]/label/span")))
                vat_radio.click()
                fix_nip(acc_input_val)

            acc_input = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "inputType")))
            acc_input.clear()
            for tx in acc_input_val:
                acc_input.send_keys(tx)
                #abc = float(config_dict["Input_delay"])
                time.sleep(float(config_dict["Input_delay"]))
            if first_search is True:
                search_button = WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.ID, "sendTwo")))
                first_search = False
            else:
                search_button = WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.ID, "sendOne")))

            search_button.click()
            try:
                search_id_element = WebDriverWait(driver, 2).until(EC.presence_of_element_located((By.XPATH,
                                "//*[@id='tableOne']/div[2]/div/div[1]/div")))
                print_button = WebDriverWait(driver, 5).until(
                    EC.presence_of_element_located((By.ID, "superPrintButton")))
                item_found = True
                status_msg = "Account found."
            except:
                #errors
                status_msg = "Incorrect account number."
            if item_found is False:
                try:
                    search_error = WebDriverWait(driver, 2).until(EC.presence_of_element_located((By.ID, "errorBox")))
                    search_id_element = WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.XPATH,
                                "//*[@id='ext']/div/div/div[1]/div")))
                    print_button = WebDriverWait(driver, 5).until(
                        EC.presence_of_element_located((By.ID, "errorPrintButton")))
                    item_found = True
                    status_msg = "Account not found."
                except:
                    status_msg = "Incorrect account number."

            if item_found is True:
                search_id_txt = search_id_element.text
                search_id_split = search_id_txt.split(":")
                search_id = search_id_split[1]
                print_button.click()
                time.sleep(2)

            wb.worksheets[0].cell(ii, search_column).value = search_id.replace(" ", "")
            wb.worksheets[0].cell(ii, search_column+1).value = status_msg
        ii += 1

    ii = 2
    while wb.worksheets[0].cell(ii, acc_column).value is not None:
        if wb.worksheets[0].cell(ii, acc_column).value != "Bank account cannot be found":
            f_name = "potwierdzenie-" + wb.worksheets[0].cell(ii, search_column).value + ".pdf"
            f_path = os.path.join(config_dict["Downloads_path"], f_name)
            d_path = os.path.join(config_dict["Destination_path"], f_name)
            if os.path.isfile(f_path):
                shutil.move(f_path, d_path)
            else:
                wb.worksheets[0].cell(ii, search_column+1).value = wb.worksheets[0].cell(ii, search_column+1).value + " File not found"
        ii += 1

    wb.save(excel_path)
    wb.close()
    driver.close()

    ctypes.windll.user32.MessageBoxW(0, "Script finished!", "Information", 1)

if __name__ == "__main__":
    main()

import sys
import os
import datetime
sys.path.append(os.path.join(os.getcwd(), 'Lib', 'site-packages'))

from openpyxl import load_workbook

class ExcelManager:
    def __init__(self, file):
        self._wb = load_workbook(filename=file, keep_vba=False, read_only=False)
        self._sheet = self._wb['GL Cari Hesap Ekstresi 100%']

    def get_row_range(self):
        """return indexes of the first and last rows in excel with data """
        self.start_row_to_index = 1
        for i, cell in enumerate(self._sheet.values, start=1):
            if cell[0] is not None:
                self.start_row_to_index = i
                break
        self.last_row_to_index = len(list(self._sheet.values))
        return self.start_row_to_index, self.last_row_to_index

    def get_list_of_invoices(self, color='00000000', col_letter1='H', col_letter2='I'):
        """return list of dicts base on excel data

            row_num = row number in excel
            po_num = number of Purchase Order in excel
            form = status of form, "var" if it's OK
            invoice = status of invoice, "var" if it's OK

        """
        self.list_ = []
        for row in self._sheet.iter_rows(min_row=self.start_row_to_index + 1, 
                                        max_col=1, 
                                        max_row=self.last_row_to_index):
            for cell in row:
                if cell.fill.start_color.index == color and cell.value is not None:
                    if self.check_status_in_xls(self._sheet[col_letter1 + str(row[0].row)].value,
                                                self._sheet[col_letter2 + str(row[0].row)].value):
                        dict_ = {
                            'row_num': str(row[0].row),
                            'po_num': str(cell.value),
                            'form': self._sheet[col_letter1 + str(row[0].row)].value,
                            'invoice': self._sheet[col_letter2 + str(row[0].row)].value
                        }
                    
                        self.list_.append(dict_)
        return self.list_

    def save(self, filename):
        return self._wb.save(filename)

    @staticmethod
    def check_status_in_xls(*args):
        print(args)
        for i in args:
            if i != 'var':
                return True

    def set_value(self, coords, new_value):
        self._sheet[coords].value = new_value

    @property
    def sheet(self):
        return self._sheet



    





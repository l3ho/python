import sys
import os
import uuid
working_directory = os.getcwd()
sys.path.append(os.path.join(working_directory, 'Lib', 'mailer'))
sys.path.append(os.path.join(working_directory, 'Lib', 'site-packages'))
from exchangelib import Credentials, Account, Configuration, DELEGATE, FileAttachment, Message,\
                        EWSTimeZone, EWSDateTime, HTMLBody
from openpyxl import load_workbook
from openpyxl import Workbook


def connect_to_mailbox(user_mailbox, user_pass, smpt_address=None):
    if not smpt_address:
        smpt_address = user_mailbox
    server = r'outlook.office365.com'
    credentials = Credentials(user_mailbox, user_pass)
    config = Configuration(server=server, credentials=credentials)
    account = Account(primary_smtp_address=smpt_address, config=config, autodiscover=False, access_type=DELEGATE)

    ews_url = account.protocol.service_endpoint
    ews_auth_type = account.protocol.auth_type
    primary_smtp_address = account.primary_smtp_address

    config = Configuration(service_endpoint=ews_url, credentials=credentials, auth_type=ews_auth_type)
    account = Account(
        primary_smtp_address=primary_smtp_address, 
        config=config, autodiscover=False, 
        access_type=DELEGATE,
    )
    return account

def convert_xls_be_to_list(row):
    result = list()
    be_values_from_xlsx = row[3:6]
    if all(be.value is None for be in be_values_from_xlsx):
        return [82, 84, 83]
    for i, be in enumerate(be_values_from_xlsx):
        if i == 0 and be.value is not None:
            result.append(82)
        elif i == 1 and be.value is not None:
            result.append(84)
        elif i == 2 and be.value is not None:
            result.append(83)
    return result

def parse_vendor_list_xlsx(filename):
    vendor_list = dict()
    wb = load_workbook(filename=filename, keep_vba=True, read_only=False)
    sheet = wb['Main']
    for row in sheet.iter_rows():
        vendor_card = {
            'name': row[0].value,
            'email': row[2].value,
            'be': convert_xls_be_to_list(row),
        }
        vendor_list[row[1].value] = vendor_card

    del vendor_list['Vendor ID']

    return vendor_list

def convert_vendor_ids_for_oracle_q(vendors):
    ids = [str(i) for i in vendors.keys()]
    return ', '.join(ids)


def create_xlsx(data, temp_folder, date):
    wb = Workbook()
    dest_filename = str(uuid.uuid4()) + '.xlsx'
    ws1 = wb.active
    ws1.title = "lines"
    ws1.append(['Numara', 'GL Tarihi', 'Odeme Tutari'])
    for row in data:
        ws1.append([row[1], date, row[0]])
    filename = os.path.join(temp_folder, dest_filename)
    wb.save(filename=filename)
    return filename



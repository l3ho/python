import sys
import os
import datetime
import time
import re
sys.path.append(os.path.join(os.getcwd(), 'Lib', 'site-packages'))

import xlrd

from openpyxl import Workbook
from openpyxl import load_workbook
from reports import get_report_data, get_unique_po_from_report, compare_suppliers, check_suppliers_for_gib, check_report_folder, user_report

start = time.time()
# ['FMS', 'FLT', 'PCSD', 'INVOICE', 'MYBUY']
parent_folder = os.path.abspath(os.path.join(os.getcwd(), os.pardir))
REPORTS_FOLDER = os.path.join(parent_folder, 'reports')
grni_files = check_report_folder(REPORTS_FOLDER)
print(grni_files)

mybuy_report = []

wb = xlrd.open_workbook(grni_files['MUBUY'])
sheet = wb.sheet_by_index(0)
for i in range(0, sheet.nrows)[1:]:
    row = sheet.row_values(rowx=i)
    if re.search(r'[A-Z]', str(row[8]).upper()) and re.search(r'\d', str(row[8])):
        try:
            invoice = re.sub(r'[^\w]', '', str(row[8]))
            supplier = re.search(r'(.*?)_', str(row[4])).groups(0)[0].upper()
            be = re.search(r'(.*)_', str(row[3])).groups(0)[0]
            kabul = str(row[11]).replace('.0', '')
            mybuy_report.append((row[3], supplier, invoice, be, kabul))
        except:
            pass

fms_report = get_report_data(grni_files['FMS'], 'Accrual- PO Base')
flt_report = get_report_data(grni_files['FLT'], 'Accrual- PO Base')
pcsd_report = get_report_data(grni_files['PCSD'], 'Accrual- PO Base')

fms_report_receipts = get_report_data(grni_files['FMS'], 'PO Receipts')
flt_report_receipts = get_report_data(grni_files['FLT'], 'PO Receipts')
pcsd_report_receipts = get_report_data(grni_files['PCSD'], 'PO Receipts')

fms_duplicates = get_unique_po_from_report(fms_report)
flt_duplicates = get_unique_po_from_report(flt_report)
pcsd_duplicates = get_unique_po_from_report(pcsd_report)

grni_reports = {
    'FLT': {
        'REPORT': flt_report,
        'DUPLICATES': flt_duplicates,
        'REPORT_2': flt_report_receipts
    },
    'FMS': {
        'REPORT': fms_report,
        'DUPLICATES': fms_duplicates,
        'REPORT_2': fms_report_receipts
    },
    'PCSD': {
        'REPORT': pcsd_report,
        'DUPLICATES': pcsd_duplicates,
        'REPORT_2': pcsd_report_receipts
    },
}

coupa_report = []

wb = xlrd.open_workbook(grni_files['INVOICE'])
sheet = wb.sheet_by_index(0)
for i in range(0, sheet.nrows):
    row = sheet.row_values(rowx=i)
    if i == 0:
        row[1:1] = ['Note', 'Unit Price', 'Kalan Miktar', 'Match Option(P/R)', 'PO Number', 'Pers', 'Satir No', 'Sevkiyat No', 'Kabul Edilen Miktar']
    else:
        row[1:1] = [''] * 9
    if str(row[0]) != '':
        if str(row[22]) != 'CQ02-QuantityVariance' and str(row[22]) != 'CQ01-PriceVariance':
            coupa_report.append(row)

# 105566, 29

for a, row in enumerate(coupa_report):
    for i in mybuy_report:
        if (str(row[0]) in str(i[2]) and i[3] in str(row[11])) or (len(str(row[20])) > 5 and str(row[20]) != '' and str(row[20]) == str(i[2]) and i[3] in str(row[11])):
            if str(row[0]).upper().startswith('GIB') and check_suppliers_for_gib(row[14], i[1]):
                continue
            if compare_suppliers(row[14], i[1], use_blank=True):
                coupa_report[a][5] = str(i[0])
                coupa_report[a][9] = str(i[4])
                break

final_report = []
for a, i in enumerate(coupa_report):
    first_option_to_match = False
    if i[5] == '':
        for be in grni_reports.keys():
            if be in str(i[11]).upper():
                for r in grni_reports[be]['REPORT']:
                    
                    if compare_suppliers(str(i[14]), str(r[1])):
                        if r[0] == 'P':
                            break
                        elif r[0] == 'R':
                            i[4] = 'R'

    i[6] = user_report.get(i[14], '')
    for be in grni_reports.keys():
        if be in str(i[5]).upper():
            first_line = True
            for r in grni_reports[be]['REPORT']:                
                if str(r[6]).strip() != '' and str(r[6]) == str(i[5]).replace(be + '_', ''):
                    first_option_to_match = True
                    i[4] = r[0]
                    if first_line:
                        i[7] = str(r[7]).replace('.0', '')
                        i[2] = str(r[13])
                        i[3] = str(r[16])

                    if str(r[18]).endswith('25800901'):
                        i[1] = 'Fixed Assets'
                    elif str(r[18]).endswith('15700114'):
                        i[1] = 'ekmaliyet (logistics)'
                    elif be == 'FLT' and re.findall(r'\.\d{3}', str(r[18])):
                        a = re.findall(r'\.\d{3}', str(r[18]))
                        if a[-1][1:] == '770' or a[-1][1:] == '760':
                            i[1] = a[-1][1:]

                    if str(r[6]) in grni_reports[be]['DUPLICATES']:                     
                        break

                    if not first_line:
                        # pass
                        new_line = i.copy()
                        new_line[2] = str(r[13])
                        new_line[3] = str(r[16])
                        new_line[7] = str(r[7]).replace('.0', '')
                        new_line[5] = str(i[5]).replace('FMS_', '').replace('FLT_', '').replace('PCSD_', '')
                        final_report.append(new_line)

                    first_line = False

    if str(i[5]) != '':
        i[5] = str(i[5]).replace('FMS_', '').replace('FLT_', '').replace('PCSD_', '')

    if first_option_to_match:
        final_report.append(i)
        continue

    global_append = True

    if not first_option_to_match and str(i[20]) != '':
        for be in grni_reports.keys():
            for r in grni_reports[be]['REPORT_2']:
                if i[20] in r[12]:
                    new_line = i.copy()
                    #match option
                    new_line[4] = r[0]
                    # po
                    new_line[5] = r[2]
                    # siparis sira no
                    new_line[7] = r[3]
                    # sevkiyat no
                    new_line[8] = r[4]
                    # unit price
                    new_line[2] = r[18]
                    # kalan miktar/ teslim alinan miktar
                    new_line[3] = r[19]
                    # print(f'{i[20]} found in {r[12]}')
                    # print('matched', be)
                    # print(i)
                    final_report.append(new_line)
                    global_append = False

    if global_append:
        final_report.append(i)

po_and_be = dict()
for i in final_report:
    if i[5].isdigit():
        if i[11] not in po_and_be.keys():
            po_and_be[i[11]] = set()
        po_and_be[i[11]].add(i[5])

po_exceptions = dict()

for k, v in grni_reports.items():
    for row in v['REPORT']:
        if 'TUR-' + k in po_and_be:
            if str(row[6]) not in po_and_be['TUR-' + k]:
                if k not in po_exceptions.keys():
                    po_exceptions[k] = list()
                row.extend([''] * 2)
                for j in grni_reports[k]['REPORT_2']:
                    if [row[1], row[6], row[12], row[13], row[14], row[15], row[16]] == [
                        j[1], j[2], j[17], j[18], j[19], j[20], j[21]]:
                        # row[5] = j[10]
                        row[-2] = j[9]
                        row[-1] = j[13]
                        break

                if row[-2] == '':
                    if row[5] != '' and row[5] != 'ARIBA, XX':
                        row[-2] = row[5]
                    else:
                        if row[-4] != '':
                            row[-2] = row[-4]
                        else:
                            row[-2] = row[-3]

                po_exceptions[k].append(row)

wb = Workbook()
dest_filename = 'output.xlsx'
ws1 = wb.active
ws1.title = "range names"

for row in final_report:
    ws1.append(row)
ws2 = wb.create_sheet('GRNI FLT')

for row in po_exceptions['FLT']:
    ws2.append(row)
ws3 = wb.create_sheet('GRNI PCSD')

if 'PCSD' in po_exceptions:
    for row in po_exceptions['PCSD']:
        ws3.append(row)
ws4 = wb.create_sheet('GRNI FMS')

for row in po_exceptions['FMS']:
    ws4.append(row)

wb.save(filename = os.path.join(REPORTS_FOLDER, dest_filename))
print(time.time() - start)



